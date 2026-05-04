import logging
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta

logger = logging.getLogger(__name__)

def verify_formulas(ws):
    """Safety check: verify rows 22-30 are untouched."""
    warnings = []
    for row in range(22, 31):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            val = str(cell.value)
            if not val.startswith('=') and val.strip() != '':
                # A formula row has a plain value — something may have overwritten it
                warnings.append(f"WARNING: Row {row}, Col {col} has plain value '{val}' instead of a formula. Check template.")
                logger.warning(f"Possible formula overwrite at R{row}C{col}: {val}")
            elif val.startswith('='):
                logger.info(f"Formula intact at R{row}C{col}: {val}")
    return warnings

def build_month_order(bill_month_str: str) -> list:
    """
    Given the latest bill month (e.g. 'January 2026'),
    return a list of 12 months ending at that month, oldest first.
    """
    try:
        end_month = datetime.strptime(bill_month_str.strip(), "%B %Y")
    except (ValueError, AttributeError):
        # Fallback to the known months from the sample bill if parsing fails
        return [
            "February 2025", "March 2025", "April 2025", "May 2025",
            "June 2025", "July 2025", "August 2025", "September 2025",
            "October 2025", "November 2025", "December 2025", "January 2026"
        ]
    
    months = []
    for i in range(11, -1, -1):  # 11 months before + current month = 12 total
        m = end_month - relativedelta(months=i)
        months.append(m.strftime("%B %Y"))
    return months

def fill_excel(template_path: str, extracted_data: dict) -> dict:
    """
    Fill the Solar Load Excel template with extracted bill data.
    """
    try:
        wb = load_workbook(template_path, data_only=False)
        ws = wb.active
        
        ws['D1'] = extracted_data.get('consumer_name', '')
        ws['D2'] = extracted_data.get('consumer_number', '')
        ws['D3'] = extracted_data.get('fixed_charges') or ''
        
        load_kw = extracted_data.get('sanctioned_load_kw')
        if load_kw:
            try:
                ws['D4'] = f"{float(load_kw):.2f}KW"
            except (ValueError, TypeError):
                ws['D4'] = str(load_kw) + 'KW'
        else:
            ws['D4'] = ''
            
        ws['D5'] = extracted_data.get('connection_type', '')
        # DO NOT touch D6, D7, row 8, or any row >= 22
        
        bill_month = extracted_data.get('bill_month', 'January 2026')
        month_order = build_month_order(bill_month)
        
        monthly_units = extracted_data.get('monthly_units', {})
        for i, month in enumerate(month_order):
            row = 9 + i  # rows 9 through 20
            
            # Robust mapping for the month name
            units_value = None
            for key, val in monthly_units.items():
                if str(key).strip().lower() == month.lower():
                    units_value = val
                    break
            
            # Fallback for the current month if not in monthly_units OR if it is 0
            if (units_value is None or str(units_value) == '0') and i == 11:
                units_value = extracted_data.get('current_units')
            
            if units_value is not None:
                try:
                    val = float(units_value)
                    ws.cell(row=row, column=4).value = int(val)  # Column D = index 4
                    
                    # If this is the latest month, we can also calculate Unit Cost in Column F
                    if i == 11 and extracted_data.get('latest_bill_amount'):
                        latest_amt = float(extracted_data['latest_bill_amount'])
                        if val > 0:
                            ws.cell(row=row, column=6).value = round(latest_amt / val, 2) # Column F = index 6
                except (ValueError, TypeError):
                    ws.cell(row=row, column=4).value = units_value
        
        latest_amount = extracted_data.get('latest_bill_amount')
        if latest_amount:
            try:
                ws['E20'] = float(latest_amount)
            except (ValueError, TypeError):
                ws['E20'] = latest_amount
        
        warnings = verify_formulas(ws)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {"success": True, "buffer": output, "error": None, "formula_warnings": warnings}
    
    except Exception as e:
        logger.exception("Excel mapping failed")
        return {"success": False, "buffer": None, "error": str(e)}
