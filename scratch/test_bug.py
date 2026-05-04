import os, json
from utils import excel_writer

# Mock data where monthly_units and current_units differ
data = {
    "bill_month": "January 2026",
    "current_units": 999, # The "buggy" value
    "monthly_units": {
        "January 2026": 1234, # The "correct" value
        "February 2025": 88,
        # ... others
    }
}

template_path = "template/solar_template.xlsx"
result = excel_writer.fill_excel(template_path, data)

if result["success"]:
    from openpyxl import load_workbook
    wb = load_workbook(result["buffer"])
    ws = wb.active
    print(f"Value at D20: {ws['D20'].value}")
else:
    print(f"Error: {result['error']}")
